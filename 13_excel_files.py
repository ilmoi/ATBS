import random
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl


print("-------------------- SHEETS --------------------")
wb = openpyxl.load_workbook("automate_online-materials/example.xlsx")
print(type(wb))

# list sheetnames
print(wb.sheetnames)

# get a single sheet
s3 = wb['Sheet3']
print(s3)
print(s3.title)

# get the active sheet
aS = wb.active
print(aS)

# find the size of the sheet
s1 = wb['Sheet1']
print(s1.max_row)
print(s1.max_column)


print("-------------------- CELLS --------------------")
# get a single cell and its details
s1 = wb['Sheet1']
a1 = s1['A1']
print(a1)
print(a1.value)
print(a1.row)
print(a1.column)
print(a1.coordinate)

# specify cells using numbers (starting with 1)
b1 = s1.cell(row=1, column=2)
print(b1)
print(b1.value)

# go through a column, print all cells
for i in range(1, 10):
    print(i, s1.cell(row=i, column=2).value)

# list all cells
for i in range(1, s1.max_column+1):
    for j in range(1, s1.max_row+1):
        print(s1[
            str(get_column_letter(i)) + str(j)
        ].value)

print("-------------------- COLS/ROWS --------------------")
# I FIND THESE CONFUSING. DONT USE THESE. USE THE ABOVE METHODS TO GET TO CELLS / ROWS / COLS.
# col = list(s1.columns)  # need to use list to be able to use indexing
# row = list(s1.rows)
# for c in col:
#     print(c)
# print(row[1])

# convert between column letters and numbers
# NOTE the complicated import at the top:)
print(get_column_letter(2))
print(get_column_letter(88))
print(get_column_letter(999))

print(column_index_from_string('B'))
print(column_index_from_string('CJ'))
print(column_index_from_string('ALK'))


print("-------------------- SLICE --------------------")
# slice the worksheet
slice = tuple(s1['A1':'C3'])
print(slice)
for row in slice:
    for cell in row:
        print(cell, cell.value)


print("-------------------- CREATE / DEL SHEETS --------------------")
# create a new blank workbook
wb = openpyxl.Workbook()
print(wb.sheetnames)  # starts with 1 sheet
# we can change it
sheet = wb.active
sheet.title = "bacon bacon spam and eggs!"
print(wb.sheetnames)
# and save it
wb.save('example_copy.xlsx')

# add new sheets
wb.create_sheet(index=0, title='im the first!')
wb.create_sheet(index=2, title='im the last!')
wb.create_sheet(index=2, title='im the middle!')
print(wb.sheetnames)

# remove sheets
del wb['im the last!']
print(wb.sheetnames)
# again need to save it at the end
wb.save('example_copy.xlsx')


print("-------------------- WRITE VALUES --------------------")
# much like working with a dictionary
sheet = wb['im the first!']
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)

for row in range(1, 10):
    for col in range(1, 10):
        sheet[str(get_column_letter(col))+str(row)] = 'randomzz textzz!'

wb.save('example_copy.xlsx')


print("-------------------- FORMATTING --------------------")
# cells
italic24Font = Font(size=24, italic=True, name='Monospace')
sheet['A1'].font = italic24Font
wb.save('example_copy.xlsx')

# rows/cols
# NOTE the values are in "points", and should be between 0 and 409. the default value is 12.75
sheet.row_dimensions[1].height = 170
# NOTE between 0 and 255. the default value is 8.43 (also how many characters can actually be displayed in a cell)
sheet.column_dimensions['B'].width = 170

# freeze
sheet.freeze_panes = 'B2'  # = freezes row 1 and col 1


print("-------------------- FORMULAS --------------------")
for i in range(1, 10):
    sheet['B'+str(i)] = random.randint(1, 100)

sheet['B11'] = '=SUM(B1:B10)'
print(sheet['B11'].value)
wb.save('example_copy.xlsx')
# the data_only flag can be used to read the actual value from the cell not the formula
wb2 = openpyxl.load_workbook('example_copy2.xlsx', data_only=True)
sheet2 = wb2.active
print(sheet2['B11'].value)


print("-------------------- UN / MERGING CELLS --------------------")
sheet.merge_cells('A5:C7')
# this range needs to match the above, eg can't specify just half the range
sheet.unmerge_cells('A5:C7')
wb.save('example_copy.xlsx')


print("-------------------- CHARTS --------------------")
# making a chart follows the following steps
# 1 Create a reference object - points to where the data lives
refObj = openpyxl.chart.Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=10)
# 2 Create a series object
seriesObj = openpyxl.chart.Series(refObj, title='A series of random bullshit!')
# 3 Create a chart object
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'some nonsense data'
# 4 Append the series object to the chart object
chartObj.append(seriesObj)
# Add the chart object to the worksheet object, optionally specifying which cell should be top left corner of the chart
sheet.add_chart(chartObj, 'G10')
wb.save('example_copy.xlsx')
