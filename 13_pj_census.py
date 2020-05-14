import openpyxl

wb = openpyxl.load_workbook('automate_online-materials/censuspopdata.xlsx')
s = wb['Population by Census Tract']


# ==============================================================================
# # THE BELOW WORKS BUT IS SUPER INEFFICIENT
# s = tuple(s['A2':'D100000']) #this is painful
# # print(s)
#
# # 1 create a counties set
# counties = set([])
# # WORKS WITH FULL WORKSHEET
# # for cell in list(s.columns)[2]:
# #     counties.add(cell.value)
# # WORKS WITH SLICES
# for row in s:
#     counties.add(list(row)[2].value)
# print(counties)
#
# # 2 create a dict containing counties and all unique censuses
# # 3 count them
# # 4 count the pop at once
# county_censuses = {}
# for county in counties:
#     censuses = set([])
#     total_pop = 0
#
#     for row in s:
#         if list(row)[2].value == county:
#             censuses.add(list(row)[0].value)
#             total_pop += list(row)[3].value
#
#     county_censuses[county] = [len(censuses), total_pop]
# print(county_censuses)


# ==============================================================================
# THE BELOW WORKS BUT IS SUPER INEFFICIENT

# how they taught me at dataquest
# really not very fucking efficient
countyData = {}
# for row in range(2, s.max_row+1):
#     state = s['B'+str(row)].value
#     county = s['C'+str(row)].value
#     pop = s['D'+str(row)].value
#
#     if state in countyData:
#         if county in countyData[state]:
#             countyData[state][county]["pop"] += int(pop)
#             countyData[state][county]["tracts"] += 1
#         else:
#             countyData[state][county] = {"pop": 0, "tracts": 0}
#             # you still need to run these 2
#             countyData[state][county]["pop"] += int(pop)
#             countyData[state][county]["tracts"] += 1
#     else:
#         countyData[state] = {}
#         # you still need to run these 2
#         if county in countyData[state]:
#             countyData[state][county]["pop"] += int(pop)
#             countyData[state][county]["tracts"] += 1
#         else:
#             countyData[state][county] = {"pop": 0, "tracts": 0}
#             countyData[state][county]["pop"] += int(pop)
#             countyData[state][county]["tracts"] += 1

# atbs way - avoids the need for messy if statements
for row in range(2, s.max_row+1):
    state = s['B'+str(row)].value
    county = s['C'+str(row)].value
    pop = s['D'+str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'pop': 0, 'tracts': 0})

    countyData[state][county]['pop'] += int(pop)
    countyData[state][county]['tracts'] += 1


with open('countyData.txt', 'w') as f:
    f.write(str(countyData))
    print('done!')
