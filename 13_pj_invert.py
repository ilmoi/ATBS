import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('produceFixed2.xlsx')
sheet1 = wb.active
wb.create_sheet(title='converted')
sheet2 = wb['converted']

for i in range(1, sheet1.max_column+1):
    temp_list = []
    for j in range(1, 1000):  # limiting to 1k as otherwise it crashes (Too many cols generated on next sheet?)
        cell = str(get_column_letter(i)) + str(j)
        # print(cell, sheet1[cell].value)
        temp_list.append(
            sheet1[
                str(get_column_letter(i)) + str(j)
            ].value
        )
    sheet2.append(temp_list)

for i in range(1, sheet2.max_column+1):
    for j in range(1, sheet2.max_row+1):
        print(sheet2[
            str(get_column_letter(i)) + str(j)
        ].value)

wb.save('produceFixed.xlsx')
