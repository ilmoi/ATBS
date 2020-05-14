import openpyxl
from openpyxl.utils import get_column_letter

number = 10

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, number+1):
    sheet['A'+str(i)] = i
    sheet[get_column_letter(i)+'1'] = i

for i in range(2, number+1):
    for j in range(2, number+1):
        sheet[get_column_letter(i)+str(j)] = f'=A{str(j)}*{get_column_letter(i)}1'

wb.save('multiplication.xlsx')
