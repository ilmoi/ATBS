import openpyxl

wb = openpyxl.load_workbook('automate_online-materials/produceSales.xlsx')
s = wb.active
print(f'size of wb is {s.max_row} x {s.max_column}')

# MY FIRST ATTEMPT
# for row in range(2, s.max_row+1):
#     produce = s['A'+str(row)].value
#     cost_per_p = s['B'+str(row)].value
#     pounds = s['C'+str(row)].value
#
#     if produce == 'Celery':
#         s['B'+str(row)] = cost_per_p = 1.19  # note the double assign! nice:)
#     elif produce == 'Garlic':
#         s['B'+str(row)] = cost_per_p = 3.07
#     elif produce == 'Lemon':
#         s['B'+str(row)] = cost_per_p = 1.27
#
#     # update total to make it correct
#     s['D'+str(row)] = float(cost_per_p) * float(pounds)  # note you have to use floats!
#
# wb.save('produceFixed.xlsx')

# A BETTER WAY
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

for row in range(2, s.max_row+1):
    produce = s['A'+str(row)].value
    cost_per_p = s['B'+str(row)].value
    pounds = s['C'+str(row)].value

    if produce in PRICE_UPDATES:
        s['B'+str(row)] = cost_per_p = PRICE_UPDATES[produce]

    # update total to make it correct
    s['D'+str(row)] = float(cost_per_p) * float(pounds)  # note you have to use floats!

wb.save('produceFixed2.xlsx')
