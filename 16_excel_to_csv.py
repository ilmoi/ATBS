"""Convert xls into csv easily. Each sheet a separate csv file."""

import openpyxl
from openpyxl.utils import get_column_letter
import csv

wb = openpyxl.load_workbook('multiplication.xlsx', data_only=True)

for sheet in wb.worksheets:
    t = sheet.title
    s = wb[t]

# s = wb['tab1']

    s_csv = open(f'multiplication_{t}.csv', 'w')
    writer = csv.writer(s_csv)

    for j in range(1, s.max_row+1):
        row_as_list = []
        for i in range(1, s.max_column+1):
            cell = s[
                str(get_column_letter(i)) + str(j)
            ].value
            row_as_list.append(cell)
        print(row_as_list)
        writer.writerow(row_as_list)

    s_csv.close()
