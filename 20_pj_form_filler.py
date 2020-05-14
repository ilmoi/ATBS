"""convert data from a excel to data in a form"""

import os
import openpyxl
import pyautogui

os.chdir('/Users/ilja/Dropbox/atbs')

# ==============================================================================
# create fake date
# wb = openpyxl.Workbook()
# s = wb.active
#
# names = ['ilja', 'vlad', 'potato']
# fears = ['bedbugs', 'bugbeds', 'shovels']
# powers = ['Want', 'Amulet', 'Discount crystall ball']
# robo = [5, 1, 3]
# comments = ['what a great book!', 'thank you for writing!', 'you da man!']
#
# row = 1
# for name, fear, power, robo_rating, comment in zip(names, fears, powers, robo, comments):
#     s['A'+str(row)] = name
#     s['B'+str(row)] = fear
#     s['C'+str(row)] = power
#     s['D'+str(row)] = robo_rating
#     s['E'+str(row)] = comment
#     row += 1
#
# wb.save('fake_users.xlsx')

# ==============================================================================
# get data from excel
wb = openpyxl.load_workbook('fake_users.xlsx')
s = wb.active
print(s)

fake_users = []
for row in range(1, s.max_row+1):
    name = s['A'+str(row)].value
    fear = s['B'+str(row)].value
    power = s['C'+str(row)].value
    robo_rating = s['D'+str(row)].value
    comment = s['E'+str(row)].value

    fake_users.append({
        "name": name,
        "fear": fear,
        "power": power,
        "robo_rating": int(robo_rating),
        "comment": comment
    })

# print(fake_users)
# print(len(fake_users))

# ==============================================================================

for fake_user in fake_users:
    # click into form: 168,389
    pyautogui.click(168, 389)  # first click to switch app
    pyautogui.click(168, 389)  # second click to target the right field
    pyautogui.sleep(1)

    # fill name
    pyautogui.write(fake_user['name'], 0.25)
    pyautogui.write(['tab'])
    pyautogui.sleep(1)

    # fill fear
    pyautogui.write(fake_user['fear'], 0.25)
    pyautogui.write(['tab'])
    pyautogui.sleep(1)

    # fill wizards
    if fake_user['power'] == 'Want':
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['return'])
    elif fake_user['power'] == 'Amulet':
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['return'])
    elif fake_user['power'] == 'Discount crystall ball':
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['return'])
    else:
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['down'])
        pyautogui.sleep(1)
        pyautogui.write(['return'])
    pyautogui.sleep(1)  # you really do need a lot of print statements...
    pyautogui.write(['tab'])
    pyautogui.sleep(1)

    # fill robocop
    for i in range(fake_user['robo_rating']-1):
        pyautogui.write(['right'])
        pyautogui.sleep(1)
    pyautogui.write(['tab'])
    pyautogui.sleep(1)

    # fill comments
    pyautogui.write(fake_user['comment'], 0.25)
    pyautogui.write(['tab'])
    pyautogui.sleep(1)

    # submit
    pyautogui.write(['return'])
    pyautogui.sleep(1)

    # reset
    pyautogui.write(['tab'])
    pyautogui.sleep(1)
    pyautogui.write(['return'])
    pyautogui.sleep(2)
